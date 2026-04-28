import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import re
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

BASE = r"c:\Users\kevin\Desktop\Emergency\Last One"

# ============================================================
# Helpers
# ============================================================

def normalize_nc(val):
    if val is None:
        return None
    s = str(val).strip().replace('\u200b', '').replace('\xa0', ' ')
    s = re.sub(r'[_\s]+$', '', s)
    m = re.match(r'(?:NC[_\s]*)?(\d+)', s, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None

def parse_euro(val):
    """Parse '568,00 €' or '1.135,00 €' into a float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = s.replace('€', '').replace('\u202f', '').replace('\xa0', '').strip()
    if not s or s.lower() in ('none', 'n/a', '-', ''):
        return None
    s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None

def parse_date(val):
    """Parse dates from various formats into a datetime.date or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s or s.lower() in ('none', 'n/a', '-'):
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def fmt_date(d):
    if d is None:
        return ''
    return d.strftime('%d/%m/%Y')

def fmt_num(n):
    if n is None:
        return ''
    if n == int(n):
        return str(int(n))
    return f"{n:.2f}"

# ============================================================
# 1. Load Senora Rosetta V5 (PMS/Rosetta data)
# ============================================================
print("Loading Senora Rosetta V5...")
wb1 = openpyxl.load_workbook(os.path.join(BASE, "Senora Rosetta V5.xlsx"), read_only=True, data_only=True)
ws1 = wb1[wb1.sheetnames[0]]
ros_rows = list(ws1.iter_rows(values_only=True))
wb1.close()

ros_header = ros_rows[0]
print(f"  Rows: {len(ros_rows)-1}, Columns: {len(ros_header)}")

rosetta_by_nc = {}
rosetta_dups = {}
for r in ros_rows[1:]:
    nc = normalize_nc(r[0])
    if nc is None:
        continue
    entry = {
        'nc': nc,
        'nc_raw': r[0],
        'contract_id': r[1],
        'deposit': r[2] if isinstance(r[2], (int, float)) else parse_euro(r[2]),
        'rent': r[3] if isinstance(r[3], (int, float)) else parse_euro(r[3]),
        'start_date': parse_date(r[4]),
        'end_date': parse_date(r[5]),
        'tenant_name': r[6],
        'tenant_email': r[7],
        'tenant_id': r[8],
        'contract_type': r[9],
        'tenancy_status': r[10],
    }
    if nc in rosetta_by_nc:
        if nc not in rosetta_dups:
            rosetta_dups[nc] = [rosetta_by_nc[nc]]
        rosetta_dups[nc].append(entry)
    rosetta_by_nc[nc] = entry

print(f"  Unique NCs: {len(rosetta_by_nc)}")
print(f"  Duplicate NCs (multiple rows): {len(rosetta_dups)}")

# ============================================================
# 2. Load Unified Contracts Data (contract extraction)
# ============================================================
print("\nLoading Unified Contracts Data v1.0...")
wb2 = openpyxl.load_workbook(os.path.join(BASE, "Unified_Contracts_Data_v1.0.xlsx"), read_only=True, data_only=True)
ws2 = wb2[wb2.sheetnames[0]]
uni_rows = list(ws2.iter_rows(values_only=True))
wb2.close()

uni_header = uni_rows[0]
print(f"  Rows: {len(uni_rows)-1}, Columns: {len(uni_header)}")

unified_by_nc = {}
unified_dups = {}
for r in uni_rows[1:]:
    nc = normalize_nc(r[0])
    if nc is None:
        continue
    entry = {
        'nc': nc,
        'nc_raw': r[0],
        'deposit': parse_euro(r[1]),
        'rent': parse_euro(r[2]),
        'start_date': parse_date(r[3]),
        'end_date': parse_date(r[4]),
    }
    if nc in unified_by_nc:
        if nc not in unified_dups:
            unified_dups[nc] = [unified_by_nc[nc]]
        unified_dups[nc].append(entry)
    unified_by_nc[nc] = entry

print(f"  Unique NCs: {len(unified_by_nc)}")
print(f"  Duplicate NCs (multiple rows): {len(unified_dups)}")

# ============================================================
# 3. Compute matching stats
# ============================================================
all_ncs = sorted(set(rosetta_by_nc.keys()) | set(unified_by_nc.keys()))
both_ncs = sorted(set(rosetta_by_nc.keys()) & set(unified_by_nc.keys()))
ros_only = sorted(set(rosetta_by_nc.keys()) - set(unified_by_nc.keys()))
uni_only = sorted(set(unified_by_nc.keys()) - set(rosetta_by_nc.keys()))

print(f"\n{'='*80}")
print("MATCHING SUMMARY")
print(f"{'='*80}")
print(f"  Total unique NCs across both:   {len(all_ncs)}")
print(f"  NCs in BOTH files:              {len(both_ncs)}")
print(f"  NCs only in Rosetta (PMS):      {len(ros_only)}")
print(f"  NCs only in Unified (Contract): {len(uni_only)}")

# Field-by-field comparison for matched NCs
deposit_match = deposit_mismatch = deposit_both_empty = 0
rent_match = rent_mismatch = rent_both_empty = 0
start_match = start_mismatch = start_both_empty = 0
end_match = end_mismatch = end_both_empty = 0

for nc in both_ncs:
    rd = rosetta_by_nc[nc]
    ud = unified_by_nc[nc]
    
    # Deposit
    rv, uv = rd['deposit'], ud['deposit']
    if rv is None and uv is None:
        deposit_both_empty += 1
    elif rv is not None and uv is not None and abs(rv - uv) < 0.01:
        deposit_match += 1
    else:
        deposit_mismatch += 1
    
    # Rent
    rv, uv = rd['rent'], ud['rent']
    if rv is None and uv is None:
        rent_both_empty += 1
    elif rv is not None and uv is not None and abs(rv - uv) < 0.01:
        rent_match += 1
    else:
        rent_mismatch += 1
    
    # Start Date
    rv, uv = rd['start_date'], ud['start_date']
    if rv is None and uv is None:
        start_both_empty += 1
    elif rv == uv:
        start_match += 1
    else:
        start_mismatch += 1
    
    # End Date
    rv, uv = rd['end_date'], ud['end_date']
    if rv is None and uv is None:
        end_both_empty += 1
    elif rv == uv:
        end_match += 1
    else:
        end_mismatch += 1

print(f"\n  Field comparison for {len(both_ncs)} matched NCs:")
print(f"  {'Field':<20} {'Match':>8} {'Mismatch':>10} {'Both Empty':>12}")
print(f"  {'-'*52}")
print(f"  {'Deposit':<20} {deposit_match:>8} {deposit_mismatch:>10} {deposit_both_empty:>12}")
print(f"  {'Rent':<20} {rent_match:>8} {rent_mismatch:>10} {rent_both_empty:>12}")
print(f"  {'Start Date':<20} {start_match:>8} {start_mismatch:>10} {start_both_empty:>12}")
print(f"  {'End Date':<20} {end_match:>8} {end_mismatch:>10} {end_both_empty:>12}")

# Show some mismatches
print(f"\n  Sample DEPOSIT mismatches (first 5):")
shown = 0
for nc in both_ncs:
    if shown >= 5:
        break
    rd = rosetta_by_nc[nc]
    ud = unified_by_nc[nc]
    rv, uv = rd['deposit'], ud['deposit']
    if not (rv is None and uv is None):
        if rv is None or uv is None or abs(rv - uv) >= 0.01:
            print(f"    NC_{nc:04d}: Rosetta={fmt_num(rv):>10}  Contract={fmt_num(uv):>10}  Name={rd.get('tenant_name','')}")
            shown += 1

print(f"\n  Sample RENT mismatches (first 5):")
shown = 0
for nc in both_ncs:
    if shown >= 5:
        break
    rd = rosetta_by_nc[nc]
    ud = unified_by_nc[nc]
    rv, uv = rd['rent'], ud['rent']
    if not (rv is None and uv is None):
        if rv is None or uv is None or abs(rv - uv) >= 0.01:
            print(f"    NC_{nc:04d}: Rosetta={fmt_num(rv):>10}  Contract={fmt_num(uv):>10}  Name={rd.get('tenant_name','')}")
            shown += 1

print(f"\n  Sample START DATE mismatches (first 5):")
shown = 0
for nc in both_ncs:
    if shown >= 5:
        break
    rd = rosetta_by_nc[nc]
    ud = unified_by_nc[nc]
    rv, uv = rd['start_date'], ud['start_date']
    if not (rv is None and uv is None) and rv != uv:
        print(f"    NC_{nc:04d}: Rosetta={fmt_date(rv):>12}  Contract={fmt_date(uv):>12}  Name={rd.get('tenant_name','')}")
        shown += 1

print(f"\n  Sample END DATE mismatches (first 5):")
shown = 0
for nc in both_ncs:
    if shown >= 5:
        break
    rd = rosetta_by_nc[nc]
    ud = unified_by_nc[nc]
    rv, uv = rd['end_date'], ud['end_date']
    if not (rv is None and uv is None) and rv != uv:
        print(f"    NC_{nc:04d}: Rosetta={fmt_date(rv):>12}  Contract={fmt_date(uv):>12}  Name={rd.get('tenant_name','')}")
        shown += 1

# ============================================================
# 4. Create the output Excel
# ============================================================
print(f"\n{'='*80}")
print("GENERATING OUTPUT FILE...")
print(f"{'='*80}")

out_wb = openpyxl.Workbook()

# -- Styles --
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill_nc = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_fill_ros = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
header_fill_con = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
header_fill_comp = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
header_fill_info = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
mismatch_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
empty_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
ros_only_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
uni_only_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ============================================================
# SHEET 1: Full Comparison
# ============================================================
ws_comp = out_wb.active
ws_comp.title = "Full Comparison"

headers = [
    ("NC", header_fill_nc),
    ("Rosetta Deposit", header_fill_ros),
    ("Contract Deposit", header_fill_con),
    ("Deposit Match?", header_fill_comp),
    ("Rosetta Rent", header_fill_ros),
    ("Contract Rent", header_fill_con),
    ("Rent Match?", header_fill_comp),
    ("Rosetta Start Date", header_fill_ros),
    ("Contract Start Date", header_fill_con),
    ("Start Date Match?", header_fill_comp),
    ("Rosetta End Date", header_fill_ros),
    ("Contract End Date", header_fill_con),
    ("End Date Match?", header_fill_comp),
    ("Overall Status", header_fill_comp),
    ("Contract ID", header_fill_info),
    ("Tenant Name", header_fill_info),
    ("Tenant Email", header_fill_info),
    ("Tenant ID", header_fill_info),
    ("Contract Type", header_fill_info),
    ("Tenancy Status", header_fill_info),
    ("Source", header_fill_info),
]

for ci, (hname, hfill) in enumerate(headers, 1):
    cell = ws_comp.cell(row=1, column=ci, value=hname)
    cell.font = header_font
    cell.fill = hfill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

ws_comp.freeze_panes = 'A2'
ws_comp.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

row_idx = 2
for nc in all_ncs:
    rd = rosetta_by_nc.get(nc)
    ud = unified_by_nc.get(nc)
    
    nc_str = f"NC_{nc:04d}"
    
    if rd and ud:
        source = "Both"
        
        # Deposit comparison
        r_dep = rd['deposit']
        u_dep = ud['deposit']
        if r_dep is None and u_dep is None:
            dep_match = "Both Empty"
            dep_fill = empty_fill
        elif r_dep is not None and u_dep is not None and abs(r_dep - u_dep) < 0.01:
            dep_match = "MATCH"
            dep_fill = match_fill
        else:
            dep_match = "MISMATCH"
            dep_fill = mismatch_fill
        
        # Rent comparison
        r_rent = rd['rent']
        u_rent = ud['rent']
        if r_rent is None and u_rent is None:
            rent_match = "Both Empty"
            rent_fill = empty_fill
        elif r_rent is not None and u_rent is not None and abs(r_rent - u_rent) < 0.01:
            rent_match = "MATCH"
            rent_fill = match_fill
        else:
            rent_match = "MISMATCH"
            rent_fill = mismatch_fill
        
        # Start date comparison
        r_sd = rd['start_date']
        u_sd = ud['start_date']
        if r_sd is None and u_sd is None:
            sd_match = "Both Empty"
            sd_fill = empty_fill
        elif r_sd == u_sd:
            sd_match = "MATCH"
            sd_fill = match_fill
        else:
            sd_match = "MISMATCH"
            sd_fill = mismatch_fill
        
        # End date comparison
        r_ed = rd['end_date']
        u_ed = ud['end_date']
        if r_ed is None and u_ed is None:
            ed_match = "Both Empty"
            ed_fill = empty_fill
        elif r_ed == u_ed:
            ed_match = "MATCH"
            ed_fill = match_fill
        else:
            ed_match = "MISMATCH"
            ed_fill = mismatch_fill
        
        statuses = [dep_match, rent_match, sd_match, ed_match]
        mismatches = sum(1 for s in statuses if s == "MISMATCH")
        empties = sum(1 for s in statuses if s == "Both Empty")
        if mismatches == 0 and empties == 0:
            overall = "ALL MATCH"
            overall_fill = match_fill
        elif mismatches == 0:
            overall = f"Match ({empties} empty)"
            overall_fill = match_fill
        else:
            overall = f"{mismatches} MISMATCH"
            overall_fill = mismatch_fill
        
        row_data = [
            nc_str,
            r_dep, u_dep, dep_match,
            r_rent, u_rent, rent_match,
            fmt_date(r_sd), fmt_date(u_sd), sd_match,
            fmt_date(r_ed), fmt_date(u_ed), ed_match,
            overall,
            rd.get('contract_id', ''),
            rd.get('tenant_name', ''),
            rd.get('tenant_email', ''),
            rd.get('tenant_id', ''),
            rd.get('contract_type', ''),
            rd.get('tenancy_status', ''),
            source,
        ]
        match_fills = [None, None, None, dep_fill, None, None, rent_fill,
                       None, None, sd_fill, None, None, ed_fill, overall_fill,
                       None, None, None, None, None, None, None]
        
    elif rd:
        source = "Rosetta Only"
        row_data = [
            nc_str,
            rd['deposit'], '', 'Rosetta Only',
            rd['rent'], '', 'Rosetta Only',
            fmt_date(rd['start_date']), '', 'Rosetta Only',
            fmt_date(rd['end_date']), '', 'Rosetta Only',
            'ROSETTA ONLY',
            rd.get('contract_id', ''),
            rd.get('tenant_name', ''),
            rd.get('tenant_email', ''),
            rd.get('tenant_id', ''),
            rd.get('contract_type', ''),
            rd.get('tenancy_status', ''),
            source,
        ]
        match_fills = [None] + [ros_only_fill]*3 + [ros_only_fill]*3 + [ros_only_fill]*3 + [ros_only_fill]*3 + [ros_only_fill] + [None]*6 + [None]
        
    else:
        source = "Contract Only"
        row_data = [
            nc_str,
            '', ud['deposit'], 'Contract Only',
            '', ud['rent'], 'Contract Only',
            '', fmt_date(ud['start_date']), 'Contract Only',
            '', fmt_date(ud['end_date']), 'Contract Only',
            'CONTRACT ONLY',
            '', '', '', '', '', '',
            source,
        ]
        match_fills = [None] + [uni_only_fill]*3 + [uni_only_fill]*3 + [uni_only_fill]*3 + [uni_only_fill]*3 + [uni_only_fill] + [None]*6 + [None]
    
    for ci, val in enumerate(row_data, 1):
        cell = ws_comp.cell(row=row_idx, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if match_fills[ci-1]:
            cell.fill = match_fills[ci-1]
    
    row_idx += 1

# Column widths
col_widths = [12, 16, 16, 14, 14, 14, 14, 18, 18, 16, 18, 18, 16, 16, 12, 35, 30, 18, 14, 14, 14]
for ci, w in enumerate(col_widths, 1):
    ws_comp.column_dimensions[get_column_letter(ci)].width = w

# ============================================================
# SHEET 2: Mismatches Only
# ============================================================
ws_mis = out_wb.create_sheet("Mismatches Only")

for ci, (hname, hfill) in enumerate(headers, 1):
    cell = ws_mis.cell(row=1, column=ci, value=hname)
    cell.font = header_font
    cell.fill = hfill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

ws_mis.freeze_panes = 'A2'

row_idx = 2
for nc in both_ncs:
    rd = rosetta_by_nc[nc]
    ud = unified_by_nc[nc]
    
    has_mismatch = False
    
    r_dep, u_dep = rd['deposit'], ud['deposit']
    dep_ok = (r_dep is None and u_dep is None) or (r_dep is not None and u_dep is not None and abs(r_dep - u_dep) < 0.01)
    
    r_rent, u_rent = rd['rent'], ud['rent']
    rent_ok = (r_rent is None and u_rent is None) or (r_rent is not None and u_rent is not None and abs(r_rent - u_rent) < 0.01)
    
    r_sd, u_sd = rd['start_date'], ud['start_date']
    sd_ok = (r_sd is None and u_sd is None) or r_sd == u_sd
    
    r_ed, u_ed = rd['end_date'], ud['end_date']
    ed_ok = (r_ed is None and u_ed is None) or r_ed == u_ed
    
    if dep_ok and rent_ok and sd_ok and ed_ok:
        continue
    
    nc_str = f"NC_{nc:04d}"
    dep_match = "MATCH" if dep_ok else ("Both Empty" if r_dep is None and u_dep is None else "MISMATCH")
    rent_match_s = "MATCH" if rent_ok else ("Both Empty" if r_rent is None and u_rent is None else "MISMATCH")
    sd_match = "MATCH" if sd_ok else ("Both Empty" if r_sd is None and u_sd is None else "MISMATCH")
    ed_match = "MATCH" if ed_ok else ("Both Empty" if r_ed is None and u_ed is None else "MISMATCH")
    
    mismatches = sum(1 for x in [dep_ok, rent_ok, sd_ok, ed_ok] if not x)
    overall = f"{mismatches} MISMATCH"
    
    row_data = [
        nc_str,
        r_dep, u_dep, dep_match,
        r_rent, u_rent, rent_match_s,
        fmt_date(r_sd), fmt_date(u_sd), sd_match,
        fmt_date(r_ed), fmt_date(u_ed), ed_match,
        overall,
        rd.get('contract_id', ''),
        rd.get('tenant_name', ''),
        rd.get('tenant_email', ''),
        rd.get('tenant_id', ''),
        rd.get('contract_type', ''),
        rd.get('tenancy_status', ''),
        "Both",
    ]
    
    fills = [
        None,
        mismatch_fill if not dep_ok else None, mismatch_fill if not dep_ok else None, mismatch_fill if not dep_ok else match_fill,
        mismatch_fill if not rent_ok else None, mismatch_fill if not rent_ok else None, mismatch_fill if not rent_ok else match_fill,
        mismatch_fill if not sd_ok else None, mismatch_fill if not sd_ok else None, mismatch_fill if not sd_ok else match_fill,
        mismatch_fill if not ed_ok else None, mismatch_fill if not ed_ok else None, mismatch_fill if not ed_ok else match_fill,
        mismatch_fill,
        None, None, None, None, None, None, None,
    ]
    
    for ci, val in enumerate(row_data, 1):
        cell = ws_mis.cell(row=row_idx, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if fills[ci-1]:
            cell.fill = fills[ci-1]
    
    row_idx += 1

for ci, w in enumerate(col_widths, 1):
    ws_mis.column_dimensions[get_column_letter(ci)].width = w

ws_mis.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ============================================================
# SHEET 3: Summary Stats
# ============================================================
ws_stats = out_wb.create_sheet("Summary")

stats_data = [
    ("Metric", "Value"),
    ("", ""),
    ("OVERVIEW", ""),
    ("Total Unique NCs (both files)", len(all_ncs)),
    ("NCs in Both Files", len(both_ncs)),
    ("NCs only in Rosetta (PMS)", len(ros_only)),
    ("NCs only in Contract Extraction", len(uni_only)),
    ("", ""),
    ("ROSETTA FILE", ""),
    ("Total Rosetta rows", len(ros_rows)-1),
    ("Rosetta rows with NC", len(rosetta_by_nc)),
    ("Rosetta duplicate NCs", len(rosetta_dups)),
    ("", ""),
    ("CONTRACT FILE", ""),
    ("Total Contract rows", len(uni_rows)-1),
    ("Contract rows with NC", len(unified_by_nc)),
    ("Contract duplicate NCs", len(unified_dups)),
    ("", ""),
    ("FIELD COMPARISON (matched NCs only)", ""),
    ("", ""),
    ("Deposit - Match", deposit_match),
    ("Deposit - Mismatch", deposit_mismatch),
    ("Deposit - Both Empty", deposit_both_empty),
    ("", ""),
    ("Rent - Match", rent_match),
    ("Rent - Mismatch", rent_mismatch),
    ("Rent - Both Empty", rent_both_empty),
    ("", ""),
    ("Start Date - Match", start_match),
    ("Start Date - Mismatch", start_mismatch),
    ("Start Date - Both Empty", start_both_empty),
    ("", ""),
    ("End Date - Match", end_match),
    ("End Date - Mismatch", end_mismatch),
    ("End Date - Both Empty", end_both_empty),
]

for ri, (metric, value) in enumerate(stats_data, 1):
    c1 = ws_stats.cell(row=ri, column=1, value=metric)
    c2 = ws_stats.cell(row=ri, column=2, value=value)
    c1.border = thin_border
    c2.border = thin_border
    if ri == 1:
        c1.font = header_font
        c1.fill = header_fill_nc
        c2.font = header_font
        c2.fill = header_fill_nc
    elif metric and metric.isupper():
        c1.font = Font(bold=True, size=11)

ws_stats.column_dimensions['A'].width = 40
ws_stats.column_dimensions['B'].width = 15

# ============================================================
# SHEET 4: Rosetta Only NCs
# ============================================================
if ros_only:
    ws_ro = out_wb.create_sheet("Rosetta Only")
    ro_headers = ["NC", "Contract ID", "Deposit", "Rent", "Start Date", "End Date",
                  "Tenant Name", "Tenant Email", "Tenant ID", "Contract Type", "Tenancy Status"]
    for ci, h in enumerate(ro_headers, 1):
        cell = ws_ro.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill_ros
        cell.border = thin_border
    
    for ri, nc in enumerate(ros_only, 2):
        rd = rosetta_by_nc[nc]
        vals = [f"NC_{nc:04d}", rd['contract_id'], rd['deposit'], rd['rent'],
                fmt_date(rd['start_date']), fmt_date(rd['end_date']),
                rd['tenant_name'], rd['tenant_email'], rd['tenant_id'],
                rd['contract_type'], rd['tenancy_status']]
        for ci, v in enumerate(vals, 1):
            cell = ws_ro.cell(row=ri, column=ci, value=v)
            cell.border = thin_border

# ============================================================
# SHEET 5: Contract Only NCs
# ============================================================
if uni_only:
    ws_uo = out_wb.create_sheet("Contract Only")
    uo_headers = ["NC", "Deposit", "Rent", "Start Date", "End Date"]
    for ci, h in enumerate(uo_headers, 1):
        cell = ws_uo.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill_con
        cell.border = thin_border
    
    for ri, nc in enumerate(uni_only, 2):
        ud = unified_by_nc[nc]
        vals = [f"NC_{nc:04d}", ud['deposit'], ud['rent'],
                fmt_date(ud['start_date']), fmt_date(ud['end_date'])]
        for ci, v in enumerate(vals, 1):
            cell = ws_uo.cell(row=ri, column=ci, value=v)
            cell.border = thin_border

# Save
outpath = os.path.join(BASE, "Rosetta_vs_Contract_Comparison.xlsx")
out_wb.save(outpath)
print(f"\nSaved to: {outpath}")
print("DONE!")
