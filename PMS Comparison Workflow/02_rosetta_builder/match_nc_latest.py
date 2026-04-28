import openpyxl
from openpyxl.styles import Font
import os
import sys
import re
from collections import defaultdict
import unicodedata

sys.stdout.reconfigure(encoding='utf-8')

BASE_NC = r"c:\Users\kevin\Desktop\Emergency\PMS Issue"
BASE_PMS = r"c:\Users\kevin\Desktop\Emergency\Last One"
PMS_FILE = os.path.join(BASE_PMS, "Tenancy-Latest2.xlsx")
OUTPUT_FILE = os.path.join(BASE_PMS, "Tenancy-Latest2_with_NC.xlsx")

# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def normalize_nc(val):
    if val is None:
        return None
    s = str(val).strip().replace('\u200b', '').replace('\xa0', ' ')
    s = re.sub(r'[_\s]+$', '', s)
    m = re.match(r'(?:NC[_\s]*)?(\d+)', s, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None

def strip_accents(s):
    nfkd = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))

def clean_name(name):
    if name is None:
        return ''
    s = str(name).strip()
    s = strip_accents(s)
    s = s.lower()
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    prefixes = ['dna ', 'd ', 'dna. ', 'don ', 'dona ']
    for p in prefixes:
        if s.startswith(p):
            s = s[len(p):]
    return s.strip()

def clean_id(id_val):
    if id_val is None:
        return None
    s = str(id_val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    s = s.upper()
    s = re.sub(r'[\s\-\.]', '', s)
    if s in ('', 'NONE', 'N/A', '-', '0', 'TEST'):
        return None
    return s

def split_tenant_ids(tid_str):
    if tid_str is None:
        return []
    s = str(tid_str).strip()
    if s.endswith('.0'):
        s = s[:-2]
    parts = re.split(r'[|,;/]', s)
    result = []
    for p in parts:
        c = clean_id(p.strip())
        if c:
            result.append(c)
    return result

def name_tokens(name):
    return set(clean_name(name).split())

def name_similarity(n1, n2):
    t1 = name_tokens(n1)
    t2 = name_tokens(n2)
    if not t1 or not t2:
        return 0.0
    intersection = t1 & t2
    if not intersection:
        return 0.0
    return len(intersection) / min(len(t1), len(t2))

# ============================================================
# STEP 1: Parse new PMS file (read-only, for matching logic only)
# ============================================================
print("STEP 1: Parsing new PMS (Tenancy-Latest2.xlsx) for matching...")
wb_pms = openpyxl.load_workbook(PMS_FILE, read_only=True, data_only=True)
ws_pms = wb_pms[wb_pms.sheetnames[0]]

pms_contracts = []
for i, row in enumerate(ws_pms.iter_rows(values_only=True)):
    if i == 0:
        continue
    contract_id_raw = row[0]
    if contract_id_raw is None:
        continue
    try:
        contract_id = int(float(str(contract_id_raw)))
    except (ValueError, TypeError):
        continue

    raw_tid = str(row[4]) if row[4] else ''
    tenant_ids = split_tenant_ids(raw_tid)
    tenant_name = clean_name(row[5])
    raw_name = str(row[5]).strip() if row[5] else ''

    pms_contracts.append({
        'contract_id': contract_id,
        'tenant_ids': tenant_ids,
        'tenant_name': tenant_name,
        'raw_name': raw_name,
    })
wb_pms.close()
print(f"  Loaded {len(pms_contracts)} contracts")

# ============================================================
# STEP 2: Parse the 3 NC reference files
# ============================================================
print("STEP 2: Parsing NC reference files...")

nc_entries = {}

# 2a. Final Contract Extractions
wb = openpyxl.load_workbook(os.path.join(BASE_NC, "Final Contract Extractions 10.04.2026.xlsx"), read_only=True, data_only=True)
ws = wb['contract_extractions_rows (2)']
count_final = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    nc = normalize_nc(row[2])
    if nc is None:
        continue
    name = clean_name(row[4])
    id_num = clean_id(row[6])
    if nc not in nc_entries:
        nc_entries[nc] = {'nc': nc, 'ids': set(), 'names': set()}
    if name:
        nc_entries[nc]['names'].add(name)
    if id_num:
        nc_entries[nc]['ids'].add(id_num)
    count_final += 1
wb.close()
print(f"  Final Extractions: {count_final} rows")

# 2b. Ana's Doc
wb = openpyxl.load_workbook(os.path.join(BASE_NC, "Tech Version - Ana_s Doc.xlsx"), read_only=True, data_only=True)
ws = wb['CLIENTES']
count_ana = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    nc = normalize_nc(row[0])
    if nc is None:
        continue
    name = clean_name(row[1])
    id_num = clean_id(row[3])
    if name and name in ('anulado', 'anulada'):
        continue
    if nc not in nc_entries:
        nc_entries[nc] = {'nc': nc, 'ids': set(), 'names': set()}
    if name:
        nc_entries[nc]['names'].add(name)
    if id_num:
        nc_entries[nc]['ids'].add(id_num)
    count_ana += 1
wb.close()
print(f"  Ana's Doc: {count_ana} rows")

# 2c. Seguimiento (Leasing Update)
wb = openpyxl.load_workbook(os.path.join(BASE_NC, "Tech Version - Doc Seguimiento 10-04 v1.0.xlsx"), read_only=True, data_only=True)
ws = wb['Leasing Update']
count_seg = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 4:
        continue
    nc = normalize_nc(row[2])
    if nc is None:
        continue
    first_name = str(row[3]).strip() if row[3] else ''
    last_name = str(row[4]).strip() if row[4] else ''
    full_name = f"{first_name} {last_name}".strip()
    name = clean_name(full_name)
    if not name:
        continue
    if nc not in nc_entries:
        nc_entries[nc] = {'nc': nc, 'ids': set(), 'names': set()}
    nc_entries[nc]['names'].add(name)
    count_seg += 1
wb.close()
print(f"  Seguimiento: {count_seg} rows")
print(f"  Total unique NCs: {len(nc_entries)}")

# ============================================================
# STEP 3: Build lookup indexes
# ============================================================
print("STEP 3: Building lookup indexes...")

nc_by_id = defaultdict(set)
for nc, data in nc_entries.items():
    for id_num in data['ids']:
        nc_by_id[id_num].add(nc)

nc_by_name = defaultdict(set)
for nc, data in nc_entries.items():
    for name in data['names']:
        if name:
            nc_by_name[name].add(nc)

# ============================================================
# STEP 4: Match contracts to NCs (group by person)
# ============================================================
print("STEP 4: Matching contracts to NCs...")

person_contracts = defaultdict(set)
person_ncs = defaultdict(set)

for c in pms_contracts:
    cid = c['contract_id']
    matched_person = None

    for tid in c['tenant_ids']:
        if tid in nc_by_id:
            matched_person = f"ID:{tid}"
            break

    if not matched_person and c['tenant_name']:
        if c['tenant_name'] in nc_by_name:
            matched_person = f"NAME:{c['tenant_name']}"

    if not matched_person and c['tenant_name']:
        best_score = 0
        best_key = None
        cn = c['tenant_name']
        cn_tokens = name_tokens(cn)
        if len(cn_tokens) >= 2:
            for nc_name in nc_by_name:
                score = name_similarity(cn, nc_name)
                if score > best_score and score >= 0.7:
                    best_score = score
                    best_key = f"FUZZY:{nc_name}"
            if best_key:
                matched_person = best_key

    if matched_person:
        person_contracts[matched_person].add(cid)

        if matched_person.startswith("ID:"):
            tid = matched_person[3:]
            for nc in nc_by_id[tid]:
                person_ncs[matched_person].add(nc)
        elif matched_person.startswith("NAME:") or matched_person.startswith("FUZZY:"):
            name_key = matched_person.split(":", 1)[1]
            for nc in nc_by_name.get(name_key, set()):
                person_ncs[matched_person].add(nc)

# Pull in NCs via other IDs of the same person
for person_key in list(person_contracts.keys()):
    if person_key.startswith("ID:"):
        tid = person_key[3:]
        for nc in list(person_ncs[person_key]):
            if nc in nc_entries:
                for other_id in nc_entries[nc]['ids']:
                    for extra_nc in nc_by_id.get(other_id, set()):
                        person_ncs[person_key].add(extra_nc)
        for c in pms_contracts:
            if c['contract_id'] in person_contracts[person_key]:
                for other_tid in c['tenant_ids']:
                    for extra_nc in nc_by_id.get(other_tid, set()):
                        person_ncs[person_key].add(extra_nc)

# ============================================================
# STEP 5: Positional matching (sort NCs and Contracts ascending)
# ============================================================
print("STEP 5: Positional matching...")

contract_to_nc = {}

for person_key in person_contracts:
    contracts_sorted = sorted(person_contracts[person_key])
    ncs_sorted = sorted(person_ncs[person_key])

    for idx, cid in enumerate(contracts_sorted):
        if idx < len(ncs_sorted):
            contract_to_nc[cid] = ncs_sorted[idx]

matched_count = len(contract_to_nc)
total_count = len(pms_contracts)
print(f"  Matched: {matched_count}/{total_count} ({matched_count/total_count*100:.1f}%)")
print(f"  Unmatched: {total_count - matched_count}")

# ============================================================
# STEP 6: Open ORIGINAL file, insert NC column, save
#         TOUCH NOTHING ELSE
# ============================================================
print("STEP 6: Inserting NC column into original file (no other changes)...")

wb_out = openpyxl.load_workbook(PMS_FILE)
for sname in wb_out.sheetnames:
    ws_out = wb_out[sname]

    ws_out.insert_cols(1)

    ws_out.cell(row=1, column=1, value="NC")
    ws_out.cell(row=1, column=1).font = Font(bold=True)
    ws_out.column_dimensions['A'].width = 12

    for row_idx in range(2, ws_out.max_row + 1):
        contract_id_cell = ws_out.cell(row=row_idx, column=2)
        cid_raw = contract_id_cell.value
        if cid_raw is None:
            continue
        try:
            cid = int(float(str(cid_raw)))
        except (ValueError, TypeError):
            continue
        nc = contract_to_nc.get(cid)
        if nc:
            ws_out.cell(row=row_idx, column=1, value=f"NC_{nc:04d}")

wb_out.save(OUTPUT_FILE)
print(f"\nOutput saved to: {OUTPUT_FILE}")

# ============================================================
# Print sample matches
# ============================================================
print("\n" + "=" * 80)
print("SAMPLE MATCHES (first 30)")
print("=" * 80)
sample_count = 0
for c in pms_contracts:
    nc = contract_to_nc.get(c['contract_id'])
    if nc:
        print(f"  NC_{nc:04d} -> Contract {c['contract_id']:>6} | {c['raw_name'][:45]}")
        sample_count += 1
        if sample_count >= 30:
            break

print("\n" + "=" * 80)
print("SAMPLE MULTI-CONTRACT PERSONS")
print("=" * 80)
shown = 0
for person_key, cids in person_contracts.items():
    if len(cids) > 1:
        ncs = sorted(person_ncs.get(person_key, set()))
        cids_sorted = sorted(cids)
        print(f"\n  Person: {person_key}")
        for idx, cid in enumerate(cids_sorted):
            nc_assigned = contract_to_nc.get(cid, "???")
            nc_label = f"NC_{nc_assigned:04d}" if isinstance(nc_assigned, int) else nc_assigned
            name = ""
            for c in pms_contracts:
                if c['contract_id'] == cid:
                    name = c['raw_name'][:35]
                    break
            print(f"    {nc_label} -> Contract {cid:>6} | {name}")
        shown += 1
        if shown >= 10:
            break

print("\n" + "=" * 80)
print("UNMATCHED (first 20)")
print("=" * 80)
um = 0
for c in pms_contracts:
    if c['contract_id'] not in contract_to_nc:
        print(f"  Contract {c['contract_id']:>6} | TIDs: {c['tenant_ids']} | {c['raw_name'][:45]}")
        um += 1
        if um >= 20:
            break

print("\nDone!")
